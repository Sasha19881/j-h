import openpyxl
from aiogram import Router, types, F
from aiogram.types import Message
from config import config
import os
import tempfile
import shutil
from datetime import datetime
import logging

router = Router()
logger = logging.getLogger(__name__)

def check_admin(user_id: int) -> bool:
    return user_id in config.ADMIN_IDS

async def create_backup():
    if os.path.exists(config.CONTACTS_FILE):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(config.BACKUP_DIR, f"backup_{timestamp}.xlsx")
        shutil.copy2(config.CONTACTS_FILE, backup_file)

def load_data(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheets_data = {}
    
    # Get the sheet names
    sheet_names = workbook.sheetnames
    
    # Load data from each sheet and remove empty rows
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            if any(row):  # Check if the row is not empty
                data.append(row)
        sheets_data[sheet_name] = data
    
    return sheets_data

@router.message(F.text == "📁 Загрузить файл")
async def request_file(message: types.Message):
    if not check_admin(message.from_user.id):
        await message.answer("❌ Доступ запрещен")
        return
    
    await message.answer(
        "📤 Отправьте Excel-файл со структурой:\n\n"
        "- Лист 'Группа А' (обязательно с русской А)\n"
        "- Лист 'Группа Б' (обязательно с русской Б)\n"
        "- Колонки: Email, WhatsApp, Telegram"
    )

@router.message(F.document & (F.document.file_name.endswith(('.xlsx', '.xls'))))
async def handle_excel(message: Message):
    if not check_admin(message.from_user.id):
        await message.answer("❌ Доступ запрещен")
        return
    
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            temp_path = tmp.name

        file = await message.bot.get_file(message.document.file_id)
        await message.bot.download_file(file.file_path, temp_path)

        await create_backup()

        # Проверяем наличие нужных листов
        excel_file = openpyxl.load_workbook(temp_path)
        sheet_names = excel_file.sheetnames
        
        required_sheets = {'Группа А', 'Группа Б'}
        missing_sheets = required_sheets - set(sheet_names)
        
        if missing_sheets:
            await message.answer(
                f"❌ В файле отсутствуют листы: {', '.join(missing_sheets)}\n"
                "Обратите внимание на точное написание:\n"
                "- 'Группа А' (русская А)\n"
                "- 'Группа Б' (русская Б)"
            )
            return

        sheets_data = load_data(temp_path)

        # Здесь можно добавить дополнительную обработку данных, если необходимо

        await message.answer("✅ Файл успешно загружен и обработан.")

    except Exception as e:
        await message.answer(
            f"❌ Ошибка: {str(e)}"
        )
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass
